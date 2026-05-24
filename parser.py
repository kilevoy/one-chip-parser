"""Парсер one-chip.ru: категории «Загрузчики» и «Редакторы» со всеми подкатегориями.

Может работать как CLI (см. main()) и как библиотека для GUI (см. run()).

Что собирается по каждому товару:
    url, category_path, name, image, description (HTML и текст),
    price, price_currency, sku, in_stock, brand,
    variations[]   — sku-варианты (sku_id, name, price, sku_number, in_stock)
    features{}     — характеристики «название: значение»
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass, field, asdict
from typing import Callable, Optional
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup

BASE = "https://one-chip.ru"
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
    ),
    "Accept-Language": "ru-RU,ru;q=0.9,en;q=0.5",
}

ROOT_CATEGORIES = {
    "zagruzchiki": (
        "Загрузчики",
        f"{BASE}/category/chip-tyuning/zagruzchiki/",
    ),
    "redaktory": (
        "Редакторы",
        f"{BASE}/category/chip-tyuning/redaktory/",
    ),
}

REQUEST_DELAY = 0.0
RETRIES = 4
TIMEOUT = 30


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------
_session = requests.Session()
_session.headers.update(HEADERS)
# По умолчанию игнорируем системный прокси (HTTP_PROXY/HTTPS_PROXY и
# Internet Settings Windows). Локальные прокси VPN/AntiDPI часто
# обрывают соединения с целевым сайтом.
_session.trust_env = False


def use_system_proxy(enable: bool) -> None:
    """Включить/выключить использование системного прокси для HTTP-сессии."""
    _session.trust_env = bool(enable)


def http_get(url: str, stop_event: Optional[threading.Event] = None) -> str:
    """GET с ретраями. Возвращает текст HTML или '' при 404."""
    last_err: Optional[Exception] = None
    for attempt in range(1, RETRIES + 1):
        if stop_event is not None and stop_event.is_set():
            raise StoppedError()
        try:
            r = _session.get(url, timeout=TIMEOUT)
            if r.status_code == 200:
                if REQUEST_DELAY:
                    time.sleep(REQUEST_DELAY)
                return r.text
            if r.status_code == 404:
                return ""
            last_err = RuntimeError(f"HTTP {r.status_code}")
        except Exception as e:  # noqa: BLE001
            last_err = e
        time.sleep(min(2 ** attempt, 10))
    raise RuntimeError(f"Не удалось получить {url}: {last_err}")


class StoppedError(Exception):
    """Выбрасывается, когда пользователь остановил парсинг."""


# ---------------------------------------------------------------------------
# Категории
# ---------------------------------------------------------------------------

def parse_category_page(html: str) -> tuple[list[str], list[str], int]:
    """Возвращает (subcategory_urls, product_urls_on_page, last_page_number)."""
    soup = BeautifulSoup(html, "lxml")

    sub_links: set[str] = set()
    for a in soup.select(
        ".е-categories-list a[href^='/category/'], "
        ".categories-list a[href^='/category/'], "
        ".e-categories-list a[href^='/category/']"
    ):
        href = a.get("href")
        if href:
            sub_links.add(urljoin(BASE, href))

    if not sub_links:
        h1 = soup.find("h1")
        if h1:
            for a in h1.find_all_next("a", href=True):
                if a.get("href", "").startswith("/category/"):
                    sub_links.add(urljoin(BASE, a["href"]))
                if a.find_parent(class_=re.compile(r"products-list")):
                    break

    products: list[str] = []
    for it in soup.select("[data-product-id]"):
        a = it.select_one("a[href]")
        if a and a.get("href"):
            products.append(urljoin(BASE, a["href"]))

    last_page = 1
    for a in soup.select("a[href*='page=']"):
        m = re.search(r"page=(\d+)", a.get("href", ""))
        if m:
            last_page = max(last_page, int(m.group(1)))

    seen = set()
    products_unique: list[str] = []
    for u in products:
        if u not in seen:
            seen.add(u)
            products_unique.append(u)
    return sorted(sub_links), products_unique, last_page


def collect_categories(
    root_url: str,
    log: Callable[[str], None] = lambda s: None,
    stop_event: Optional[threading.Event] = None,
) -> list[str]:
    """BFS по всем подкатегориям."""
    seen: set[str] = set()
    order: list[str] = []
    queue: list[str] = [root_url]
    while queue:
        if stop_event is not None and stop_event.is_set():
            raise StoppedError()
        url = queue.pop(0)
        if url in seen:
            continue
        seen.add(url)
        order.append(url)
        try:
            html = http_get(url, stop_event=stop_event)
        except StoppedError:
            raise
        except Exception as e:  # noqa: BLE001
            log(f"  ! ошибка категории {url}: {e}")
            continue
        subs, _, _ = parse_category_page(html)
        for s in subs:
            if s.startswith(root_url) and s not in seen:
                queue.append(s)
    return order


def collect_products_in_category(
    cat_url: str,
    log: Callable[[str], None] = lambda s: None,
    stop_event: Optional[threading.Event] = None,
) -> list[str]:
    """Все ссылки на товары в категории (с пагинацией)."""
    all_products: list[str] = []
    seen: set[str] = set()
    page = 1
    while True:
        if stop_event is not None and stop_event.is_set():
            raise StoppedError()
        url = cat_url if page == 1 else f"{cat_url}?page={page}"
        try:
            html = http_get(url, stop_event=stop_event)
        except StoppedError:
            raise
        except Exception as e:  # noqa: BLE001
            log(f"  ! ошибка стр. {url}: {e}")
            break
        if not html:
            break
        _, prods, last = parse_category_page(html)
        new_added = 0
        for p in prods:
            if p not in seen:
                seen.add(p)
                all_products.append(p)
                new_added += 1
        if page >= last or new_added == 0:
            break
        page += 1
    return all_products


# ---------------------------------------------------------------------------
# Карточка товара
# ---------------------------------------------------------------------------

@dataclass
class Variation:
    sku_id: str = ""
    name: str = ""
    price: str = ""
    sku_number: str = ""
    in_stock: bool = True


@dataclass
class Product:
    url: str = ""
    discovered_in: str = ""
    discovered_root: str = ""
    category_path: str = ""
    name: str = ""
    image: str = ""
    description_text: str = ""
    description_html: str = ""
    price: str = ""
    price_text: str = ""
    price_currency: str = "RUB"
    sku: str = ""
    in_stock: bool = True
    brand: str = ""
    variations: list[Variation] = field(default_factory=list)
    features: dict[str, str] = field(default_factory=dict)

    def to_dict(self) -> dict:
        return asdict(self)


_PRICE_RE = re.compile(r"[^\d]")


def _clean_price(text: str) -> str:
    if not text:
        return ""
    return _PRICE_RE.sub("", text)


def parse_product(
    url: str,
    stop_event: Optional[threading.Event] = None,
    discovered_in: str = "",
    discovered_root: str = "",
) -> Optional[Product]:
    html = http_get(url, stop_event=stop_event)
    if not html:
        return None
    soup = BeautifulSoup(html, "lxml")
    p = Product(url=url, discovered_in=discovered_in, discovered_root=discovered_root)

    h1 = soup.find("h1")
    if h1:
        p.name = h1.get_text(" ", strip=True)

    og = soup.find("meta", property="og:image")
    if og and og.get("content"):
        p.image = og["content"]
    else:
        img = soup.select_one(".js-product-image, .e-product-image img, img.e-product__image")
        if img and (img.get("src") or img.get("data-src")):
            p.image = img.get("data-src") or img.get("src", "")

    desc = soup.select_one("[itemprop='description']")
    if desc:
        p.description_text = desc.get_text("\n", strip=True)
        p.description_html = desc.decode_contents().strip()
    # fallback: блок .e-product-description без заголовка «Описание»
    if not p.description_text:
        alt = soup.select_one(".e-product-description")
        if alt:
            # уберём заголовок «Описание», который рендерится отдельным элементом
            for h in alt.select(".e-product-description__title, h2, h3"):
                h.decompose()
            p.description_text = alt.get_text("\n", strip=True)
            p.description_html = alt.decode_contents().strip()

    price_el = soup.select_one(".e-product-prices__general .price, .e-product-prices__general")
    if price_el:
        raw = price_el.get_text(" ", strip=True)
        p.price_text = raw
        p.price = _clean_price(raw)
    cur = soup.select_one("meta[itemprop='priceCurrency']")
    if cur and cur.get("content"):
        p.price_currency = cur["content"]

    sku_el = soup.select_one(".e-product-sku")
    if sku_el and "_empty" not in (sku_el.get("class") or []):
        txt = sku_el.get_text(" ", strip=True)
        m = re.search(r"Артикул[:\s]*([^\s]+)", txt)
        if m:
            p.sku = m.group(1)
        else:
            p.sku = txt.replace("Артикул:", "").strip()

    stock_el = soup.select_one("[class*='-stock'], .e-stock, .stock")
    if stock_el:
        t = stock_el.get_text(" ", strip=True).lower()
        p.in_stock = "нет" not in t

    crumbs: list[str] = []
    for bl in soup.select("[itemtype*='BreadcrumbList']"):
        for li in bl.select("li[itemprop='itemListElement']"):
            nm = li.find(attrs={"itemprop": "name"})
            if nm:
                t = nm.get_text(" ", strip=True)
                if t and t not in ("-", "Главная", "Home"):
                    crumbs.append(t)
        if crumbs:
            break
    if crumbs and p.name and crumbs[-1] == p.name:
        crumbs = crumbs[:-1]
    p.category_path = " / ".join(crumbs)

    for li in soup.select("ul.skus li.skus__li, li.skus__li"):
        v = Variation()
        inp = li.select_one("input[name='sku_id']")
        if inp:
            v.sku_id = inp.get("value", "")
            v.price = inp.get("data-price", "") or v.price
            v.sku_number = inp.get("data-sku-number", "")
        nm = li.select_one("[itemprop='name']")
        if nm:
            v.name = nm.get_text(" ", strip=True)
        if not v.price:
            pr = li.select_one("meta[itemprop='price']")
            if pr and pr.get("content"):
                v.price = pr["content"]
        avail = li.select_one("[itemprop='availability']")
        if avail:
            href = avail.get("href", "")
            v.in_stock = "InStock" in href
        p.variations.append(v)

    for tr in soup.select(".e-product-features__table tr.e-product-features__tr"):
        name_el = tr.select_one(".e-product-features__name")
        val_el = tr.select_one(".e-product-features__value")
        if name_el and val_el:
            k = name_el.get_text(" ", strip=True)
            v = val_el.get_text(" ", strip=True)
            if k:
                p.features[k] = v
    if "Бренд" in p.features:
        p.brand = p.features["Бренд"]

    return p


# ---------------------------------------------------------------------------
# Сохранение
# ---------------------------------------------------------------------------

def _unique_path(path: str) -> str:
    """Если файл занят/существует и его нельзя перезаписать — подбираем
    свободное имя вида name_1.ext, name_2.ext, ..."""
    if not os.path.exists(path):
        return path
    # пытаемся понять, можем ли перезаписать
    try:
        with open(path, "ab"):
            pass
        return path
    except OSError:
        pass
    base, ext = os.path.splitext(path)
    i = 1
    while True:
        candidate = f"{base}_{i}{ext}"
        if not os.path.exists(candidate):
            return candidate
        try:
            with open(candidate, "ab"):
                pass
            return candidate
        except OSError:
            i += 1


def save_json(products: list[Product], path: str) -> str:
    path = _unique_path(path)
    data = [p.to_dict() for p in products]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return path


def save_csv(products: list[Product], path: str) -> str:
    """Один общий CSV. Плоский вид: одна вариация = одна строка."""
    path = _unique_path(path)
    rows = _flat_rows(products)
    headers_keys = [k for k, _ in FLAT_HEADERS]
    headers_rus = [r for _, r in FLAT_HEADERS]
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(headers_rus)
        for r in rows:
            w.writerow([r.get(k, "") for k in headers_keys])
    return path


def _flat_rows(products: list[Product]) -> list[dict]:
    """Преобразует список товаров в плоские строки: одна вариация = одна строка.
    Если вариаций нет, делается одна строка с пустым variant_name.
    """
    rows: list[dict] = []
    for p in products:
        # вычислим subcategory как последний элемент пути
        subcat = ""
        if p.category_path:
            parts = [x.strip() for x in p.category_path.split("/") if x.strip()]
            if parts:
                subcat = parts[-1]
        base_price_num = int(p.price) if p.price.isdigit() else ""
        base = {
            "main_category": p.discovered_root,
            "category_path": p.category_path,
            "subcategory": subcat,
            "product_name": p.name,
            "variant_name": "",
            "price": p.price_text or (p.price + " ₽" if p.price else ""),
            "variant_price": base_price_num,
            "product_url": p.url,
            "image_url": p.image,
            "description": p.description_text,
            "characteristics": "\n".join(f"{k}: {v}" for k, v in p.features.items()),
            "brand": p.brand,
            "sku": p.sku,
            "availability": "да" if p.in_stock else "нет",
            "source_category_url": p.discovered_in,
        }
        if not p.variations:
            base["variant_name"] = "Без вариации"
            rows.append(base)
            continue
        for v in p.variations:
            row = dict(base)
            row["variant_name"] = v.name or "Без названия"
            try:
                row["variant_price"] = int(v.price) if v.price and str(v.price).isdigit() else v.price
            except Exception:  # noqa: BLE001
                row["variant_price"] = v.price
            row["sku"] = v.sku_number or p.sku
            row["availability"] = "да" if v.in_stock else "нет"
            rows.append(row)
    return rows


FLAT_HEADERS = [
    ("main_category",       "Категория"),
    ("category_path",       "Путь категории"),
    ("subcategory",         "Подкатегория"),
    ("product_name",        "Название"),
    ("variant_name",        "Вариация"),
    ("price",               "Цена (текст)"),
    ("variant_price",       "Цена вариации"),
    ("product_url",         "Ссылка"),
    ("image_url",           "Изображение"),
    ("description",         "Описание"),
    ("characteristics",     "Характеристики"),
    ("brand",               "Бренд"),
    ("sku",                 "Артикул"),
    ("availability",        "В наличии"),
    ("source_category_url", "Найдено в подкатегории"),
]


def save_xlsx(products: list[Product], path: str) -> str:
    """Сохраняет XLSX в плоском виде: одна вариация = одна строка.
    Делает листы по корневой категории + лист «Все»."""
    path = _unique_path(path)
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, Alignment, PatternFill
    except Exception:  # noqa: BLE001
        return path

    rows = _flat_rows(products)
    wb = Workbook()
    wb.remove(wb.active)

    headers_keys = [k for k, _ in FLAT_HEADERS]
    headers_rus = [r for _, r in FLAT_HEADERS]
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="EEEEEE")
    widths = {
        "main_category": 14, "category_path": 35, "subcategory": 22,
        "product_name": 35, "variant_name": 25, "price": 14,
        "variant_price": 14, "product_url": 35, "image_url": 35,
        "description": 60, "characteristics": 40, "brand": 14,
        "sku": 14, "availability": 10, "source_category_url": 35,
    }

    def _make_sheet(title: str, items: list[dict]) -> None:
        safe = re.sub(r"[\\/?*\[\]:]", "_", title)[:31] or "Sheet"
        ws = wb.create_sheet(title=safe)
        ws.append(headers_rus)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        for r in items:
            ws.append([r.get(k, "") for k in headers_keys])
        for i, k in enumerate(headers_keys, start=1):
            ws.column_dimensions[get_column_letter(i)].width = widths.get(k, 20)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    by_root: dict[str, list[dict]] = {}
    for r in rows:
        by_root.setdefault(r.get("main_category") or "Без корня", []).append(r)
    for root_name in sorted(by_root):
        _make_sheet(root_name, by_root[root_name])
    _make_sheet("Все", rows)

    wb.save(path)
    return path


def save_csv_by_root(products: list[Product], out_dir: str) -> list[str]:
    """Один CSV на каждую корневую категорию (плоский вид)."""
    rows = _flat_rows(products)
    headers_keys = [k for k, _ in FLAT_HEADERS]
    headers_rus = [r for _, r in FLAT_HEADERS]
    by_root: dict[str, list[dict]] = {}
    for r in rows:
        by_root.setdefault(r.get("main_category") or "Без корня", []).append(r)
    paths: list[str] = []
    for root_name, items in by_root.items():
        slug = re.sub(r"[^a-zA-Zа-яА-Я0-9]+", "_", root_name).strip("_") or "root"
        path = _unique_path(os.path.join(out_dir, f"products_{slug}.csv"))
        with open(path, "w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f, delimiter=";")
            w.writerow(headers_rus)
            for r in items:
                w.writerow([r.get(k, "") for k in headers_keys])
        paths.append(path)
    return paths


# ---------------------------------------------------------------------------
# Главный сценарий с колбэками для GUI
# ---------------------------------------------------------------------------

@dataclass
class RunResult:
    products: list[Product]
    json_path: str
    csv_path: str
    xlsx_path: str


def run(
    *,
    threads: int = 8,
    out_dir: str = ".",
    targets: Optional[list[str]] = None,
    use_proxy: bool = False,
    log: Callable[[str], None] = print,
    on_progress: Callable[[int, int], None] = lambda d, t: None,
    on_phase: Callable[[str], None] = lambda s: None,
    stop_event: Optional[threading.Event] = None,
) -> RunResult:
    """Полный цикл: сбор категорий -> ссылок -> карточек -> сохранение."""
    use_system_proxy(use_proxy)
    if not targets:
        targets = list(ROOT_CATEGORIES.keys())

    os.makedirs(out_dir, exist_ok=True)

    # лог-файл рядом с результатами
    log_path = os.path.join(out_dir, "parse.log")
    try:
        log_file = open(log_path, "w", encoding="utf-8")
    except Exception:  # noqa: BLE001
        log_file = None

    _orig_log = log

    def log(s: str) -> None:  # type: ignore[no-redef]
        try:
            _orig_log(s)
        except Exception:  # noqa: BLE001
            pass
        if log_file is not None:
            try:
                log_file.write(s.rstrip() + "\n")
                log_file.flush()
            except Exception:  # noqa: BLE001
                pass

    log(f"--- Запуск {time.strftime('%Y-%m-%d %H:%M:%S')} ---")
    log(f"targets={targets} threads={threads} use_proxy={use_proxy} out_dir={out_dir}")

    # 1) Категории: соберём подкатегории каждого корня и сохраним привязку
    on_phase("Сбор категорий")
    root_map: dict[str, str] = {}  # cat_url -> человекочитаемое имя корня
    for key in targets:
        title, root = ROOT_CATEGORIES[key]
        log(f"Сбор подкатегорий: {title} -> {root}")
        cats = collect_categories(root, log=log, stop_event=stop_event)
        log(f"  найдено категорий: {len(cats)}")
        for c in cats:
            root_map.setdefault(c, title)

    # 2) Ссылки на товары + откуда взяли
    on_phase("Сбор ссылок на товары")
    log("\nСбор ссылок на товары...")
    product_urls: list[str] = []
    seen_urls: set[str] = set()
    discovery: dict[str, tuple[str, str]] = {}  # url -> (discovered_in, discovered_root)
    cats_in_order: list[str] = list(root_map.keys())
    for cat in cats_in_order:
        if stop_event is not None and stop_event.is_set():
            raise StoppedError()
        urls = collect_products_in_category(cat, log=log, stop_event=stop_event)
        added = 0
        for u in urls:
            if u not in seen_urls:
                seen_urls.add(u)
                product_urls.append(u)
                discovery[u] = (cat, root_map[cat])
                added += 1
        log(f"  {cat} -> {len(urls)} товаров (+{added} новых)")
    log(f"\nИтого уникальных товаров: {len(product_urls)}")

    # 3) Карточки
    on_phase("Парсинг карточек")
    products: list[Product] = []
    failed: list[tuple[str, str]] = []  # (url, error)
    total = len(product_urls)
    on_progress(0, total)

    def _job(u: str) -> tuple[str, Optional[Product], Optional[str]]:
        try:
            di, dr = discovery.get(u, ("", ""))
            return u, parse_product(u, stop_event=stop_event,
                                    discovered_in=di, discovered_root=dr), None
        except StoppedError:
            raise
        except Exception as e:  # noqa: BLE001
            return u, None, str(e)

    def _finalize(reason: str) -> RunResult:
        on_phase(f"Сохранение ({reason})")
        products.sort(key=lambda p: (p.discovered_root, p.category_path, p.name))
        json_path = os.path.join(out_dir, "products.json")
        csv_path = os.path.join(out_dir, "products.csv")
        xlsx_path = os.path.join(out_dir, "products.xlsx")
        failed_path = os.path.join(out_dir, "failed.txt")
        json_path = save_json(products, json_path)
        csv_path = save_csv(products, csv_path)
        try:
            xlsx_path = save_xlsx(products, xlsx_path)
        except Exception as e:  # noqa: BLE001
            log(f"  ! не удалось сохранить XLSX: {e}")
            xlsx_path = ""
        # отдельные CSV по корневой категории
        try:
            csv_per_root = save_csv_by_root(products, out_dir)
        except Exception as e:  # noqa: BLE001
            log(f"  ! не удалось сохранить CSV по категориям: {e}")
            csv_per_root = []
        # список не получившихся ссылок — пригодится для повтора
        try:
            with open(failed_path, "w", encoding="utf-8") as f:
                for url, err in failed:
                    f.write(url + "\t" + err.replace("\n", " ") + "\n")
        except Exception:  # noqa: BLE001
            pass
        log(f"\n{reason.capitalize()}:")
        log(f"  спарсено: {len(products)} | не удалось: {len(failed)}")
        log(f"  JSON: {json_path}\n  CSV : {csv_path}\n  XLSX: {xlsx_path}")
        for cp in csv_per_root:
            log(f"  CSV (категория): {cp}")
        log(f"  Список FAIL: {failed_path}")
        if log_file is not None:
            try:
                log_file.close()
            except Exception:  # noqa: BLE001
                pass
        return RunResult(products, json_path, csv_path, xlsx_path)

    try:
        with ThreadPoolExecutor(max_workers=threads) as ex:
            futures = {ex.submit(_job, u): u for u in product_urls}
            done = 0
            for fut in as_completed(futures):
                if stop_event is not None and stop_event.is_set():
                    for f in futures:
                        f.cancel()
                    raise StoppedError()
                done += 1
                try:
                    url, p, err = fut.result()
                except StoppedError:
                    raise
                except Exception as e:  # noqa: BLE001
                    failed.append((futures[fut], str(e)))
                    log(f"  [{done}/{total}] FAIL {futures[fut]}: {e}")
                    on_progress(done, total)
                    continue
                if p:
                    products.append(p)
                    log(f"  [{done}/{total}] OK {p.name[:80]}")
                elif err:
                    failed.append((url, err))
                    log(f"  [{done}/{total}] FAIL {url}: {err}")
                else:
                    failed.append((url, "пустой ответ"))
                    log(f"  [{done}/{total}] пусто {url}")
                on_progress(done, total)
    except StoppedError:
        # сохраняем то, что успели
        return _finalize("остановлено")

    return _finalize("готово")


def retry_failed(
    *,
    failed_path: str,
    out_dir: str,
    threads: int = 8,
    use_proxy: bool = False,
    log: Callable[[str], None] = print,
    on_progress: Callable[[int, int], None] = lambda d, t: None,
    on_phase: Callable[[str], None] = lambda s: None,
    stop_event: Optional[threading.Event] = None,
) -> RunResult:
    """Догрузка ранее упавших ссылок и слияние с существующим products.json."""
    use_system_proxy(use_proxy)
    on_phase("Догрузка failed.txt")

    # читаем список url из failed.txt (формат: url\\tошибка)
    urls: list[str] = []
    with open(failed_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            u = line.split("\t", 1)[0].strip()
            if u:
                urls.append(u)
    log(f"Ссылок к повторной загрузке: {len(urls)}")

    # читаем уже сохранённые товары
    products_path = os.path.join(out_dir, "products.json")
    existing: list[Product] = []
    existing_urls: set[str] = set()
    if os.path.exists(products_path):
        try:
            with open(products_path, "r", encoding="utf-8") as f:
                data = json.load(f)
            for d in data:
                vars_ = [Variation(**v) for v in d.get("variations") or []]
                p = Product(
                    url=d.get("url", ""),
                    discovered_in=d.get("discovered_in", ""),
                    discovered_root=d.get("discovered_root", ""),
                    category_path=d.get("category_path", ""),
                    name=d.get("name", ""),
                    image=d.get("image", ""),
                    description_text=d.get("description_text", ""),
                    description_html=d.get("description_html", ""),
                    price=d.get("price", ""),
                    price_text=d.get("price_text", ""),
                    price_currency=d.get("price_currency", "RUB"),
                    sku=d.get("sku", ""),
                    in_stock=d.get("in_stock", True),
                    brand=d.get("brand", ""),
                    variations=vars_,
                    features=d.get("features", {}) or {},
                )
                existing.append(p)
                existing_urls.add(p.url)
            log(f"В products.json уже есть {len(existing)} товаров")
        except Exception as e:  # noqa: BLE001
            log(f"  ! не удалось прочитать существующий products.json: {e}")

    # отфильтруем те, что уже есть
    todo = [u for u in urls if u not in existing_urls]
    log(f"Будем парсить: {len(todo)}")

    on_phase("Парсинг догруженных карточек")
    on_progress(0, len(todo))
    new_failed: list[tuple[str, str]] = []
    new_products: list[Product] = []

    def _job(u: str) -> tuple[str, Optional[Product], Optional[str]]:
        try:
            return u, parse_product(u, stop_event=stop_event), None
        except StoppedError:
            raise
        except Exception as e:  # noqa: BLE001
            return u, None, str(e)

    try:
        with ThreadPoolExecutor(max_workers=threads) as ex:
            futures = {ex.submit(_job, u): u for u in todo}
            done = 0
            total = len(todo)
            for fut in as_completed(futures):
                if stop_event is not None and stop_event.is_set():
                    for f in futures:
                        f.cancel()
                    raise StoppedError()
                done += 1
                try:
                    url, p, err = fut.result()
                except StoppedError:
                    raise
                except Exception as e:  # noqa: BLE001
                    new_failed.append((futures[fut], str(e)))
                    log(f"  [{done}/{total}] FAIL {futures[fut]}: {e}")
                    on_progress(done, total)
                    continue
                if p:
                    new_products.append(p)
                    log(f"  [{done}/{total}] OK {p.name[:80]}")
                else:
                    new_failed.append((url, err or "пустой ответ"))
                    log(f"  [{done}/{total}] FAIL {url}: {err or 'пустой ответ'}")
                on_progress(done, total)
    except StoppedError:
        log("--- остановлено ---")

    # объединяем
    all_products = list(existing) + new_products
    seen = set()
    unique: list[Product] = []
    for p in all_products:
        if p.url not in seen:
            seen.add(p.url)
            unique.append(p)
    unique.sort(key=lambda p: (p.discovered_root, p.category_path, p.name))

    on_phase("Сохранение")
    json_path = save_json(unique, os.path.join(out_dir, "products.json"))
    csv_path = save_csv(unique, os.path.join(out_dir, "products.csv"))
    try:
        xlsx_path = save_xlsx(unique, os.path.join(out_dir, "products.xlsx"))
    except Exception as e:  # noqa: BLE001
        log(f"  ! не удалось сохранить XLSX: {e}")
        xlsx_path = ""
    try:
        csv_per_root = save_csv_by_root(unique, out_dir)
    except Exception as e:  # noqa: BLE001
        log(f"  ! не удалось сохранить CSV по категориям: {e}")
        csv_per_root = []

    # обновим failed.txt: оставим только те, что снова упали
    try:
        with open(failed_path, "w", encoding="utf-8") as f:
            for url, err in new_failed:
                f.write(url + "\t" + err.replace("\n", " ") + "\n")
    except Exception:  # noqa: BLE001
        pass

    log(
        f"\nГотово.\n  Догружено: {len(new_products)} | повторно упало: {len(new_failed)}"
        f"\n  Всего в файлах: {len(unique)}"
        f"\n  JSON: {json_path}\n  CSV : {csv_path}\n  XLSX: {xlsx_path}"
    )
    for cp in csv_per_root:
        log(f"  CSV (категория): {cp}")
    on_phase("Готово")
    return RunResult(unique, json_path, csv_path, xlsx_path)
    if log_file is not None:
        try:
            log_file.close()
        except Exception:  # noqa: BLE001
            pass
    return RunResult(products, json_path, csv_path, xlsx_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--threads", type=int, default=8)
    ap.add_argument("--out-dir", default=".")
    ap.add_argument(
        "--only", choices=list(ROOT_CATEGORIES), default=None,
        help="спарсить только одну корневую категорию",
    )
    ap.add_argument(
        "--use-proxy", action="store_true",
        help="использовать системный прокси (по умолчанию игнорируется)",
    )
    args = ap.parse_args()
    targets = [args.only] if args.only else None

    def _log(s: str) -> None:
        try:
            print(s)
        except UnicodeEncodeError:
            print(s.encode("utf-8", "replace").decode(sys.stdout.encoding or "utf-8", "replace"))

    run(threads=args.threads, out_dir=args.out_dir, targets=targets,
        use_proxy=args.use_proxy, log=_log)


if __name__ == "__main__":
    main()
